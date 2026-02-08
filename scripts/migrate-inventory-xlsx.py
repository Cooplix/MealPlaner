#!/usr/bin/env python3
import argparse
import os
import sys
from datetime import datetime, time

import openpyxl
from pymongo import MongoClient


def normalize_datetime(value):
  if value is None:
    return None
  if isinstance(value, datetime):
    return value
  try:
    if hasattr(value, 'year') and hasattr(value, 'month') and hasattr(value, 'day'):
      return datetime.combine(value, time(0, 0))
  except Exception:
    pass
  return None


def main():
  parser = argparse.ArgumentParser(description='Migrate Inventory sheet into MongoDB inventory_items collection.')
  parser.add_argument('--xlsx', required=True, help='Path to inventory_system.xlsx')
  parser.add_argument('--mongo-uri', default=os.getenv('MONGODB_URI', 'mongodb://localhost:27017/'))
  parser.add_argument('--db', default=os.getenv('MONGODB_DB', 'mealplanner'))
  parser.add_argument('--user-id', required=True, help='Target user_id for inserted items')
  parser.add_argument('--dry-run', action='store_true', help='Do not write to DB')
  parser.add_argument('--clear-existing', action='store_true', help='Delete existing inventory_items for this user before insert')
  args = parser.parse_args()

  wb = openpyxl.load_workbook(args.xlsx, data_only=True)
  if 'Inventory' not in wb.sheetnames:
    print('Inventory sheet not found', file=sys.stderr)
    return 2
  ws = wb['Inventory']

  col = {
    'name': 1,
    'unit': 3,
    'quantity': 4,
    'notes': 6,
    'base_name': 7,
    'category': 8,
    'location': 9,
    'min_qty': 10,
    'max_qty': 11,
    'expires_at': 13,
  }

  items = []
  now = datetime.utcnow()
  for r in range(2, ws.max_row + 1):
    name = ws.cell(r, col['name']).value
    base_name = ws.cell(r, col['base_name']).value
    if not name and not base_name:
      continue

    item = {
      'user_id': args.user_id,
      'name': name if name is not None else base_name,
      'base_name': base_name if base_name is not None else name,
      'category': ws.cell(r, col['category']).value,
      'location': ws.cell(r, col['location']).value,
      'unit': ws.cell(r, col['unit']).value,
      'quantity': ws.cell(r, col['quantity']).value or 0,
      'min_qty': ws.cell(r, col['min_qty']).value,
      'max_qty': ws.cell(r, col['max_qty']).value,
      'expires_at': normalize_datetime(ws.cell(r, col['expires_at']).value),
      'notes': ws.cell(r, col['notes']).value,
      'added_at': now,
    }

    items.append(item)

  print(f'Prepared {len(items)} inventory items')
  if args.dry_run:
    return 0

  client = MongoClient(args.mongo_uri)
  db = client[args.db]
  coll = db['inventory_items']

  if args.clear_existing:
    deleted = coll.delete_many({'user_id': args.user_id})
    print(f'Deleted {deleted.deleted_count} existing items for user {args.user_id}')

  if items:
    coll.insert_many(items)
    print('Inserted items:', len(items))

  return 0


if __name__ == '__main__':
  raise SystemExit(main())
